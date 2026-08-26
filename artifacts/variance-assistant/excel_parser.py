"""Small, explicit helpers for reading consolidated Excel workbooks."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _column_letter(index: int) -> str:
    """Convert a zero-based column index into the familiar Excel letter."""
    letters = ""
    current = index + 1
    while current:
        current, remainder = divmod(current - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _display_value(value: Any) -> str:
    """Turn an Excel cell into a compact string for previews and form values."""
    if value is None:
        return ""
    return str(value).strip()


def _open_workbook(path: Path):
    """Open a workbook read-only so large consolidated files stay inexpensive."""
    return load_workbook(path, read_only=True, data_only=True)


def get_workbook_metadata(path: Path) -> dict[str, Any]:
    """Return sheet names, column choices, and a preview for the first sheet."""
    workbook = _open_workbook(path)
    try:
        sheets = workbook.sheetnames
        if not sheets:
            return {"sheets": [], "columns": [], "preview_rows": []}
        columns, preview_rows = get_sheet_preview(path, sheets[0])
        return {"sheets": sheets, "columns": columns, "preview_rows": preview_rows}
    finally:
        workbook.close()


def get_sheet_preview(
    path: Path, sheet_name: str, preview_limit: int = 5
) -> tuple[list[dict[str, Any]], list[list[str]]]:
    """Read the header row and a few data rows for the mapping screen."""
    workbook = _open_workbook(path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("That worksheet could not be found.")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        if header_values is None:
            raise ValueError("That worksheet is empty.")

        columns = [
            {
                "index": index,
                "name": _display_value(value) or f"Column {index + 1}",
                "label": f"{_column_letter(index)} — {_display_value(value) or f'Column {index + 1}'}",
            }
            for index, value in enumerate(header_values)
        ]
        preview_rows = [
            [_display_value(value) for value in row[: len(columns)]]
            for row in list(rows)[:preview_limit]
        ]
        return columns, preview_rows
    finally:
        workbook.close()


def extract_variance_rows(
    path: Path,
    sheet_name: str,
    category_index: int,
    forecast_index: int,
    actual_index: int,
    notes_index: int | None = None,
) -> list[dict[str, str]]:
    """Extract mapped Excel rows into the same shape as the manual input form."""
    workbook = _open_workbook(path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("That worksheet could not be found.")
        sheet = workbook[sheet_name]
        all_rows = sheet.iter_rows(values_only=True)
        next(all_rows, None)  # The first row is the selected header row.
        extracted: list[dict[str, str]] = []
        for row in all_rows:
            values = list(row)
            required_indices = [category_index, forecast_index, actual_index]
            if any(index >= len(values) for index in required_indices):
                continue
            category = _display_value(values[category_index])
            forecast = _display_value(values[forecast_index])
            actual = _display_value(values[actual_index])
            notes = (
                _display_value(values[notes_index])
                if notes_index is not None and notes_index < len(values)
                else ""
            )
            if not any((category, forecast, actual, notes)):
                continue
            extracted.append(
                {
                    "category": category,
                    "forecast": forecast,
                    "actual": actual,
                    "notes": notes,
                }
            )
        return extracted
    finally:
        workbook.close()